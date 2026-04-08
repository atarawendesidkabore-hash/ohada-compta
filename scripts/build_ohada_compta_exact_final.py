"""
build_ohada_compta_exact_final.py
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
Builds OHADA-COMPTA-EXACT.xlsm using the hosted-browser approach (no WebView2
DLL required). The output file is fully self-contained and does not break when
copied, moved, or sent to another machine — as long as Chrome or Edge is installed
on the recipient's computer and macros are enabled.

Usage:
    python scripts/build_ohada_compta_exact_final.py
    python scripts/build_ohada_compta_exact_final.py --output C:\\path\\to\\OHADA-COMPTA-EXACT.xlsm
"""

from __future__ import annotations

import argparse
import shutil
import zipfile
from pathlib import Path

import pythoncom
import win32com.client as win32
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE_WORKBOOK = ROOT / "_vendor" / "ExcelWebView2" / "ExcelWebView2.xlsm"

DEFAULT_OUTPUT = ROOT / "OHADA-COMPTA-EXACT.xlsm"
DEFAULT_DESKTOP_OUTPUT = Path.home() / "OneDrive" / "Desktop" / "OHADA-COMPTA-EXACT.xlsm"
INSTALL_PACK_DIR = Path.home() / "OneDrive" / "Desktop" / "OHADA-COMPTA-EXACT-INSTALL-PACK"
INSTALL_PACK_ZIP = Path.home() / "OneDrive" / "Desktop" / "OHADA-COMPTA-EXACT-INSTALL-PACK.zip"
INSTALL_README = INSTALL_PACK_DIR / "README-INSTALL.txt"
INSTALL_CMD = INSTALL_PACK_DIR / "INSTALL-OHADA-COMPTA-EXACT.cmd"
INSTALL_PS1 = INSTALL_PACK_DIR / "install-ohada-compta-exact.ps1"
UNINSTALL_CMD = INSTALL_PACK_DIR / "UNINSTALL-OHADA-COMPTA-EXACT.cmd"
UNINSTALL_PS1 = INSTALL_PACK_DIR / "uninstall-ohada-compta-exact.ps1"
FALLBACK_XLSX = INSTALL_PACK_DIR / "OHADA-COMPTA-EXACT-WEB-LINK.xlsx"

LIVE_SITE_URL = "https://atarawendesidkabore-hash.github.io/ohada-compta/?v=20260408a"
HOST_WINDOW_TITLE = "OHADA_COMPTA_EDGE_HOST"
USERFORM_WINDOW_TITLE = "OHADA_COMPTA_WINDOW"

# Components from the template that are not needed in the hosted-browser build
UNUSED_COMPONENTS = [
    "AppConstants",
    "APIFunctions",
    "WV2Globals",
    "MemoryFunctions",
    "clsWebViewEventHandlers",
    "clsWebViewScriptCompleteHandler",
    "clsWebViewContentHandler",
    "pluginLoader",
    "pluginExampleCls",
    "factory",
    "wv2",
    "wv2Environment",
]

INSTALL_README_TEXT = f"""OHADA COMPTA EXACT - PACK D'INSTALLATION

Ce pack installe la version Excel exacte de OHADA Compta sur le poste client.

Contenu:
- OHADA-COMPTA-EXACT.xlsm
- OHADA-COMPTA-EXACT-WEB-LINK.xlsx
- INSTALL-OHADA-COMPTA-EXACT.cmd
- UNINSTALL-OHADA-COMPTA-EXACT.cmd

Procedure conseillee:
1. Envoyez ce pack au format ZIP.
2. Le client fait clic droit sur le ZIP > Proprietes > Debloquer > Appliquer.
3. Il extrait le ZIP.
4. Il lance INSTALL-OHADA-COMPTA-EXACT.cmd.
5. Le script copie le classeur dans %LOCALAPPDATA%\\OHADA-Compta-Exact.
6. Le script cree un Trusted Location Excel et un raccourci Bureau.
7. Le client ouvre ensuite le raccourci "OHADA Compta Exact".

Ce que cela apporte:
- plus besoin d'ouvrir directement le .xlsm depuis Downloads, email ou WhatsApp
- meilleur comportement face au blocage Microsoft des macros Internet
- aucun WebView2Loader.dll ni TLB a livrer avec ce classeur

Limites:
- ce n'est pas une signature numerique editeur
- sur certains postes tres verrouilles, le service informatique peut encore devoir valider l'installation
- si les macros restent interdites, utilisez la version de secours:
  OHADA-COMPTA-EXACT-WEB-LINK.xlsx

Site direct:
{LIVE_SITE_URL}
"""

INSTALL_CMD_TEXT = """@echo off
setlocal
cd /d "%~dp0"
powershell -NoProfile -ExecutionPolicy Bypass -File "%~dp0install-ohada-compta-exact.ps1"
if errorlevel 1 (
  echo.
  echo Installation terminee avec erreur.
  pause
  exit /b 1
)
echo.
echo Installation terminee.
pause
"""

UNINSTALL_CMD_TEXT = """@echo off
setlocal
cd /d "%~dp0"
powershell -NoProfile -ExecutionPolicy Bypass -File "%~dp0uninstall-ohada-compta-exact.ps1"
if errorlevel 1 (
  echo.
  echo Desinstallation terminee avec erreur.
  pause
  exit /b 1
)
echo.
echo Desinstallation terminee.
pause
"""

INSTALL_PS1_TEXT = r"""
$ErrorActionPreference = 'Stop'

$sourceRoot = Split-Path -Parent $MyInvocation.MyCommand.Path
$installRoot = Join-Path $env:LOCALAPPDATA 'OHADA-Compta-Exact'
$macroWorkbook = 'OHADA-COMPTA-EXACT.xlsm'
$fallbackWorkbook = 'OHADA-COMPTA-EXACT-WEB-LINK.xlsx'
$macroSource = Join-Path $sourceRoot $macroWorkbook
$fallbackSource = Join-Path $sourceRoot $fallbackWorkbook
$macroTarget = Join-Path $installRoot $macroWorkbook
$fallbackTarget = Join-Path $installRoot $fallbackWorkbook

if (-not (Test-Path $macroSource)) {
    throw "Fichier introuvable: $macroSource"
}

New-Item -ItemType Directory -Path $installRoot -Force | Out-Null
Copy-Item -Path $macroSource -Destination $macroTarget -Force
if (Test-Path $fallbackSource) {
    Copy-Item -Path $fallbackSource -Destination $fallbackTarget -Force
}

Get-ChildItem -Path $installRoot -File -ErrorAction SilentlyContinue | Unblock-File -ErrorAction SilentlyContinue

$trustedPathValue = $installRoot
if (-not $trustedPathValue.EndsWith('\')) {
    $trustedPathValue += '\'
}

foreach ($version in @('16.0', '15.0', '14.0')) {
    $trustedKey = "HKCU:\Software\Microsoft\Office\$version\Excel\Security\Trusted Locations\OHADAComptaExact"
    New-Item -Path $trustedKey -Force | Out-Null
    New-ItemProperty -Path $trustedKey -Name 'Path' -PropertyType String -Value $trustedPathValue -Force | Out-Null
    New-ItemProperty -Path $trustedKey -Name 'AllowSubfolders' -PropertyType DWord -Value 1 -Force | Out-Null
    New-ItemProperty -Path $trustedKey -Name 'Description' -PropertyType String -Value 'OHADA Compta Exact' -Force | Out-Null
    New-ItemProperty -Path $trustedKey -Name 'Date' -PropertyType String -Value (Get-Date -Format 'yyyy-MM-dd HH:mm:ss') -Force | Out-Null
}

$desktop = [Environment]::GetFolderPath('Desktop')
$shortcutPath = Join-Path $desktop 'OHADA Compta Exact.lnk'
$shell = New-Object -ComObject WScript.Shell
$shortcut = $shell.CreateShortcut($shortcutPath)
$shortcut.TargetPath = $macroTarget
$shortcut.WorkingDirectory = $installRoot
$shortcut.Description = 'OHADA Compta Exact'
$shortcut.IconLocation = "$env:SystemRoot\System32\shell32.dll,1"
$shortcut.Save()

Write-Host "INSTALL_ROOT=$installRoot"
Write-Host "SHORTCUT=$shortcutPath"
Write-Host "STATUS=OK"
""".strip() + "\n"

UNINSTALL_PS1_TEXT = r"""
$ErrorActionPreference = 'Stop'

$installRoot = Join-Path $env:LOCALAPPDATA 'OHADA-Compta-Exact'
$desktop = [Environment]::GetFolderPath('Desktop')
$shortcutPath = Join-Path $desktop 'OHADA Compta Exact.lnk'

foreach ($version in @('16.0', '15.0', '14.0')) {
    $trustedKey = "HKCU:\Software\Microsoft\Office\$version\Excel\Security\Trusted Locations\OHADAComptaExact"
    if (Test-Path $trustedKey) {
        Remove-Item -Path $trustedKey -Recurse -Force
    }
}

if (Test-Path $shortcutPath) {
    Remove-Item -Path $shortcutPath -Force
}

if (Test-Path $installRoot) {
    Remove-Item -Path $installRoot -Recurse -Force
}

Write-Host "STATUS=REMOVED"
""".strip() + "\n"


# ---------------------------------------------------------------------------
# VBA source strings
# ---------------------------------------------------------------------------

HOST_MODULE_CODE = f'''
Option Explicit

Private Const GWL_STYLE As Long = -16
Private Const WS_CAPTION As Long = &HC00000
Private Const WS_SYSMENU As Long = &H80000
Private Const WS_THICKFRAME As Long = &H40000
Private Const WS_CHILD As Long = &H40000000
Private Const WS_POPUP As Long = &H80000000
Private Const WS_VISIBLE As Long = &H10000000
Private Const SWP_NOSIZE As Long = &H1
Private Const SWP_NOMOVE As Long = &H2
Private Const SWP_NOZORDER As Long = &H4
Private Const SWP_FRAMECHANGED As Long = &H20
Private Const SWP_SHOWWINDOW As Long = &H40
Private Const SW_SHOW As Long = 5
Private Const WM_CLOSE As Long = &H10
Private Const HOST_TITLE As String = "{HOST_WINDOW_TITLE}"
Private Const APP_URL As String = "{LIVE_SITE_URL}"

#If VBA7 Then
Private Declare PtrSafe Function GetWindowLongPtr Lib "user32" Alias "GetWindowLongPtrA" (ByVal hwnd As LongPtr, ByVal nIndex As Long) As LongPtr
Private Declare PtrSafe Function SetWindowLongPtr Lib "user32" Alias "SetWindowLongPtrA" (ByVal hwnd As LongPtr, ByVal nIndex As Long, ByVal dwNewLong As LongPtr) As LongPtr
Private Declare PtrSafe Function SetParent Lib "user32" (ByVal hWndChild As LongPtr, ByVal hWndNewParent As LongPtr) As LongPtr
Private Declare PtrSafe Function IsWindow Lib "user32" (ByVal hwnd As LongPtr) As Long
Private Declare PtrSafe Function IsWindowVisible Lib "user32" (ByVal hwnd As LongPtr) As Long
Private Declare PtrSafe Function MoveWindow Lib "user32" (ByVal hwnd As LongPtr, ByVal X As Long, ByVal Y As Long, ByVal nWidth As Long, ByVal nHeight As Long, ByVal bRepaint As Long) As Long
Private Declare PtrSafe Function ShowWindow Lib "user32" (ByVal hwnd As LongPtr, ByVal nCmdShow As Long) As Long
Private Declare PtrSafe Function PostMessage Lib "user32" Alias "PostMessageA" (ByVal hwnd As LongPtr, ByVal wMsg As Long, ByVal wParam As LongPtr, ByVal lParam As LongPtr) As Long
Private Declare PtrSafe Function FindWindow Lib "user32" Alias "FindWindowA" (ByVal lpClassName As String, ByVal lpWindowName As String) As LongPtr
Private Declare PtrSafe Function SetWindowPos Lib "user32" (ByVal hwnd As LongPtr, ByVal hWndInsertAfter As LongPtr, ByVal X As Long, ByVal Y As Long, ByVal cx As Long, ByVal cy As Long, ByVal uFlags As Long) As Long
Private Declare PtrSafe Sub Sleep Lib "kernel32" (ByVal dwMilliseconds As Long)
#Else
Private Declare Function GetWindowLongPtr Lib "user32" Alias "GetWindowLongA" (ByVal hwnd As Long, ByVal nIndex As Long) As Long
Private Declare Function SetWindowLongPtr Lib "user32" Alias "SetWindowLongA" (ByVal hwnd As Long, ByVal nIndex As Long, ByVal dwNewLong As Long) As Long
Private Declare Function SetParent Lib "user32" (ByVal hWndChild As Long, ByVal hWndNewParent As Long) As Long
Private Declare Function IsWindow Lib "user32" (ByVal hwnd As Long) As Long
Private Declare Function IsWindowVisible Lib "user32" (ByVal hwnd As Long) As Long
Private Declare Function MoveWindow Lib "user32" (ByVal hwnd As Long, ByVal X As Long, ByVal Y As Long, ByVal nWidth As Long, ByVal nHeight As Long, ByVal bRepaint As Long) As Long
Private Declare Function ShowWindow Lib "user32" (ByVal hwnd As Long, ByVal nCmdShow As Long) As Long
Private Declare Function PostMessage Lib "user32" Alias "PostMessageA" (ByVal hwnd As Long, ByVal wMsg As Long, ByVal wParam As Long, ByVal lParam As Long) As Long
Private Declare Function FindWindow Lib "user32" Alias "FindWindowA" (ByVal lpClassName As String, ByVal lpWindowName As String) As Long
Private Declare Function SetWindowPos Lib "user32" (ByVal hwnd As Long, ByVal hWndInsertAfter As Long, ByVal X As Long, ByVal Y As Long, ByVal cx As Long, ByVal cy As Long, ByVal uFlags As Long) As Long
Private Declare Sub Sleep Lib "kernel32" (ByVal dwMilliseconds As Long)
#End If

Private m_shellPrepared As Boolean
Private m_browserProcessId As Long
Private m_browserWindowHwnd As LongPtr

Public Sub LaunchOhadaCompta()
    On Error GoTo HandleError
    ResetOhadaTrace
    TraceOhada "LaunchOhadaCompta:start"
    PrepareExcelShell
    If UserForm1.Visible = False Then
        UserForm1.Show vbModeless
    End If
    UserForm1.Repaint
    ResizeHostedWindow
    TraceOhada "LaunchOhadaCompta:complete"
    Exit Sub
HandleError:
    TraceOhada "LaunchOhadaCompta:error:" & Err.Number & ":" & Err.Description
End Sub

Public Sub PrepareExcelShell()
    On Error Resume Next
    Application.WindowState = xlMaximized
    Application.Caption = "OHADA COMPTA"
    Application.DisplayFormulaBar = False
    Application.ExecuteExcel4Macro "SHOW.TOOLBAR(""Ribbon"",False)"
    If Not ActiveWindow Is Nothing Then
        ActiveWindow.DisplayWorkbookTabs = False
        ActiveWindow.DisplayHeadings = False
        ActiveWindow.DisplayGridlines = False
        ActiveWindow.DisplayHorizontalScrollBar = False
        ActiveWindow.DisplayVerticalScrollBar = False
    End If
    m_shellPrepared = True
End Sub

Public Sub RestoreExcelShell()
    On Error Resume Next
    StopHostedBrowser
    If m_shellPrepared Then
        Application.Caption = vbNullString
        Application.DisplayFormulaBar = True
        Application.ExecuteExcel4Macro "SHOW.TOOLBAR(""Ribbon"",True)"
        If Not ActiveWindow Is Nothing Then
            ActiveWindow.DisplayWorkbookTabs = True
            ActiveWindow.DisplayHeadings = True
            ActiveWindow.DisplayGridlines = True
            ActiveWindow.DisplayHorizontalScrollBar = True
            ActiveWindow.DisplayVerticalScrollBar = True
        End If
    End If
    m_shellPrepared = False
End Sub

Public Function ResolveBrowserExe() As String
    Dim candidate As Variant
    Dim candidates As Variant
    candidates = Array( _
        Environ$("ProgramFiles") & "\\Google\\Chrome\\Application\\chrome.exe", _
        Environ$("ProgramFiles(x86)") & "\\Google\\Chrome\\Application\\chrome.exe", _
        Environ$("LOCALAPPDATA") & "\\Google\\Chrome\\Application\\chrome.exe", _
        Environ$("ProgramFiles(x86)") & "\\Microsoft\\Edge\\Application\\msedge.exe", _
        Environ$("ProgramFiles") & "\\Microsoft\\Edge\\Application\\msedge.exe")
    For Each candidate In candidates
        If Len(candidate) > 0 Then
            If Dir$(CStr(candidate)) <> vbNullString Then
                ResolveBrowserExe = CStr(candidate)
                Exit Function
            End If
        End If
    Next candidate
End Function

Public Function HostRootPath() As String
    Dim rootPath As String
    rootPath = Environ$("LOCALAPPDATA")
    If Len(rootPath) = 0 Then rootPath = Environ$("TEMP")
    If Len(rootPath) = 0 Then rootPath = ThisWorkbook.Path
    HostRootPath = rootPath & "\\OHADA-Compta\\ShareHost"
End Function

Public Function HostHtmlPath() As String
    EnsureFolderTree HostRootPath
    HostHtmlPath = HostRootPath & "\\ohada-edge-host.html"
End Function

Public Function BrowserProfilePath() As String
    EnsureFolderTree HostRootPath
    BrowserProfilePath = HostRootPath & "\\BrowserProfile"
End Function

Public Sub EnsureFolderTree(ByVal folderPath As String)
    Dim fso As Object
    Dim parentPath As String
    If Len(folderPath) = 0 Then Exit Sub
    Set fso = CreateObject("Scripting.FileSystemObject")
    If fso.FolderExists(folderPath) Then Exit Sub
    parentPath = fso.GetParentFolderName(folderPath)
    If Len(parentPath) > 0 And Not fso.FolderExists(parentPath) Then
        EnsureFolderTree parentPath
    End If
    If Not fso.FolderExists(folderPath) Then
        fso.CreateFolder folderPath
    End If
End Sub

Public Function BuildHostHtml() As String
    BuildHostHtml = "<!doctype html>" & vbCrLf & _
        "<html lang=""en"">" & vbCrLf & _
        "<head>" & vbCrLf & _
        "  <meta charset=""utf-8"">" & vbCrLf & _
        "  <meta name=""viewport"" content=""width=device-width, initial-scale=1"">" & vbCrLf & _
        "  <title>" & HOST_TITLE & "</title>" & vbCrLf & _
        "  <style>" & vbCrLf & _
        "    html, body {{ margin: 0; width: 100%; height: 100%; overflow: hidden; background: #061426; }}" & vbCrLf & _
        "    iframe {{ width: 100%; height: 100%; border: 0; display: block; background: #061426; }}" & vbCrLf & _
        "  </style>" & vbCrLf & _
        "</head>" & vbCrLf & _
        "<body>" & vbCrLf & _
        "  <iframe src=""" & APP_URL & """ allow=""clipboard-read; clipboard-write; fullscreen""></iframe>" & vbCrLf & _
        "</body>" & vbCrLf & _
        "</html>"
End Function

Public Sub EnsureHostHtml()
    Dim fileNo As Integer
    fileNo = FreeFile
    Open HostHtmlPath For Output As #fileNo
    Print #fileNo, BuildHostHtml
    Close #fileNo
End Sub

Public Function HostHtmlUri() As String
    EnsureHostHtml
    HostHtmlUri = "file:///" & Replace(HostHtmlPath, "\\", "/")
End Function

Public Sub StartHostedBrowser()
    Dim browserExe As String
    Dim formHwnd As LongPtr
    Dim commandLine As String
    On Error GoTo HandleError
    TraceOhada "StartHostedBrowser:start"
    formHwnd = FindWindow(vbNullString, "{USERFORM_WINDOW_TITLE}")
    If formHwnd = 0 Then Err.Raise 5, "pluginExample.StartHostedBrowser", "UserForm handle unavailable"
    If m_browserWindowHwnd <> 0 Then
        If IsWindow(m_browserWindowHwnd) <> 0 Then
            AttachHostedWindow formHwnd, m_browserWindowHwnd
            ResizeHostedWindow
            TraceOhada "StartHostedBrowser:reuse"
            Exit Sub
        End If
    End If
    browserExe = ResolveBrowserExe()
    If Len(browserExe) = 0 Then Err.Raise 53, "pluginExample.StartHostedBrowser", "Chrome or Edge is not installed"
    EnsureHostHtml
    StopHostedBrowser
    commandLine = """" & browserExe & """ --app=""" & HostHtmlUri & """ --user-data-dir=""" & BrowserProfilePath & """ --new-window --disable-session-crashed-bubble --disable-features=msEdgeSidebarV2"
    TraceOhada "StartHostedBrowser:exe=" & browserExe
    m_browserProcessId = Shell(commandLine, vbNormalFocus)
    TraceOhada "StartHostedBrowser:pid=" & m_browserProcessId
    m_browserWindowHwnd = WaitForHostedWindow(15000)
    TraceOhada "StartHostedBrowser:hwnd=" & CStr(m_browserWindowHwnd)
    If m_browserWindowHwnd = 0 Then Err.Raise 5, "pluginExample.StartHostedBrowser", "Hosted browser window not found"
    AttachHostedWindow formHwnd, m_browserWindowHwnd
    ResizeHostedWindow
    TraceOhada "StartHostedBrowser:complete"
    Exit Sub
HandleError:
    TraceOhada "StartHostedBrowser:error:" & Err.Number & ":" & Err.Description
End Sub

Public Sub StopHostedBrowser()
    On Error Resume Next
    If m_browserWindowHwnd <> 0 Then
        If IsWindow(m_browserWindowHwnd) <> 0 Then
            PostMessage m_browserWindowHwnd, WM_CLOSE, 0, 0
            Sleep 300
        End If
    End If
    Shell "cmd /c taskkill /FI ""WINDOWTITLE eq " & HOST_TITLE & "*"" /T /F", vbHide
    If m_browserProcessId <> 0 Then
        Shell "cmd /c taskkill /PID " & CStr(m_browserProcessId) & " /T /F", vbHide
    End If
    m_browserWindowHwnd = 0
    m_browserProcessId = 0
End Sub

Public Sub ResizeHostedWindow()
    If m_browserWindowHwnd = 0 Then Exit Sub
    If IsWindow(m_browserWindowHwnd) = 0 Then Exit Sub
    MoveWindow m_browserWindowHwnd, 0, 0, CLng(UserForm1.InsideWidth), CLng(UserForm1.InsideHeight), 1
    ShowWindow m_browserWindowHwnd, SW_SHOW
End Sub

Private Sub AttachHostedWindow(ByVal parentHwnd As LongPtr, ByVal childHwnd As LongPtr)
    Dim style As LongPtr
    If parentHwnd = 0 Or childHwnd = 0 Then Exit Sub
    style = GetWindowLongPtr(childHwnd, GWL_STYLE)
    style = style And Not WS_CAPTION
    style = style And Not WS_SYSMENU
    style = style And Not WS_THICKFRAME
    style = style And Not WS_POPUP
    style = style Or WS_CHILD
    style = style Or WS_VISIBLE
    SetParent childHwnd, parentHwnd
    SetWindowLongPtr childHwnd, GWL_STYLE, style
    SetWindowPos childHwnd, 0, 0, 0, CLng(UserForm1.InsideWidth), CLng(UserForm1.InsideHeight), SWP_NOZORDER Or SWP_FRAMECHANGED Or SWP_SHOWWINDOW
    ShowWindow childHwnd, SW_SHOW
End Sub

Private Function WaitForHostedWindow(ByVal timeoutMs As Long) As LongPtr
    Dim startedAt As Double
    Dim hwnd As LongPtr
    startedAt = Timer
    Do
        hwnd = FindWindow(vbNullString, HOST_TITLE)
        If hwnd <> 0 Then
            If IsWindowVisible(hwnd) <> 0 Then
                WaitForHostedWindow = hwnd
                Exit Function
            End If
        End If
        If m_browserWindowHwnd <> 0 Then
            If IsWindow(m_browserWindowHwnd) <> 0 Then
                WaitForHostedWindow = m_browserWindowHwnd
                Exit Function
            End If
        End If
        DoEvents
        Sleep 100
        If Timer < startedAt Then startedAt = Timer
    Loop While ((Timer - startedAt) * 1000#) < timeoutMs
End Function

Public Sub MakeWindowFrameless(ByVal formCaption As String)
    Dim hwnd As LongPtr
    Dim style As LongPtr
    On Error Resume Next
    hwnd = FindWindow(vbNullString, formCaption)
    If hwnd = 0 Then Exit Sub
    style = GetWindowLongPtr(hwnd, GWL_STYLE)
    style = style And Not WS_CAPTION
    style = style And Not WS_SYSMENU
    style = style And Not WS_THICKFRAME
    SetWindowLongPtr hwnd, GWL_STYLE, style
    SetWindowPos hwnd, 0, 0, 0, 0, 0, SWP_NOSIZE Or SWP_NOMOVE Or SWP_NOZORDER Or SWP_FRAMECHANGED
End Sub

Public Sub ResetOhadaTrace()
    Dim fileNo As Integer
    fileNo = FreeFile
    Open TracePath For Output As #fileNo
    Print #fileNo, Format$(Now, "yyyy-mm-dd hh:nn:ss") & " | trace reset"
    Close #fileNo
End Sub

Public Sub TraceOhada(ByVal message As String)
    Dim fileNo As Integer
    fileNo = FreeFile
    Open TracePath For Append As #fileNo
    Print #fileNo, Format$(Now, "yyyy-mm-dd hh:nn:ss") & " | " & message
    Close #fileNo
End Sub

Public Function TracePath() As String
    TracePath = Environ$("TEMP") & "\\OHADA_COMPTA_SHARE_TRACE.log"
End Function
'''.strip() + "\n"


USERFORM1_CODE = f"""
Option Explicit

Private m_appModeReady As Boolean
Private m_browserStarted As Boolean

Private Sub UserForm_Initialize()
    On Error Resume Next
    ConfigureOhadaAppWindow
End Sub

Private Sub UserForm_Activate()
    On Error GoTo HandleError
    ConfigureOhadaAppWindow
    If Not m_browserStarted Then
        DoEvents
        Me.Repaint
        pluginExample.StartHostedBrowser
        m_browserStarted = True
    Else
        pluginExample.ResizeHostedWindow
    End If
    Exit Sub
HandleError:
    pluginExample.TraceOhada "UserForm_Activate:error:" & Err.Number & ":" & Err.Description
End Sub

Private Sub UserForm_Resize()
    On Error Resume Next
    ConfigureOhadaAppWindow
    pluginExample.ResizeHostedWindow
End Sub

Private Sub UserForm_QueryClose(Cancel As Integer, CloseMode As Integer)
    On Error Resume Next
    pluginExample.RestoreExcelShell
    Unload Me
End Sub

Private Sub ConfigureOhadaAppWindow()
    On Error Resume Next
    If Not m_appModeReady Then
        Me.Caption = "{USERFORM_WINDOW_TITLE}"
        cmdBack.Visible = False
        cmdForward.Visible = False
        txtUrl.Visible = False
        cmdNewTab.Visible = False
        cmdStopReload.Visible = False
        CommandButton10.Visible = False
        CommandButton7.Visible = False
        browserTabs.Visible = False
        m_appModeReady = True
    End If
    Me.StartUpPosition = 0
    Me.Left = 0
    Me.Top = 0
    Me.Width = Application.UsableWidth
    Me.Height = Application.UsableHeight
    pluginExample.MakeWindowFrameless Me.Caption
    pluginExample.ResizeHostedWindow
End Sub
""".strip() + "\n"


THISWORKBOOK_CODE = """
Option Explicit

Private Sub Workbook_Open()
    On Error GoTo HandleError
    pluginExample.LaunchOhadaCompta
    Exit Sub
HandleError:
    pluginExample.TraceOhada "Workbook_Open:error:" & Err.Number & ":" & Err.Description
End Sub

Private Sub Workbook_BeforeClose(Cancel As Boolean)
    On Error Resume Next
    pluginExample.RestoreExcelShell
End Sub
""".strip() + "\n"


# ---------------------------------------------------------------------------
# Build helpers
# ---------------------------------------------------------------------------

def remove_reference_by_name(vb_project, name: str) -> None:
    to_remove = []
    for reference in vb_project.References:
        try:
            if reference.Name == name:
                to_remove.append(reference)
        except Exception:
            pass
    for reference in to_remove:
        try:
            vb_project.References.Remove(reference)
        except Exception:
            pass


def remove_component_if_exists(vb_project, name: str) -> None:
    try:
        component = vb_project.VBComponents(name)
    except Exception:
        return
    try:
        vb_project.VBComponents.Remove(component)
    except Exception:
        module = component.CodeModule
        if module.CountOfLines:
            module.DeleteLines(1, module.CountOfLines)


def ensure_standard_module(vb_project, name: str):
    try:
        component = vb_project.VBComponents(name)
    except Exception:
        component = vb_project.VBComponents.Add(1)
        component.Name = name
    return component


def replace_component_code(component, code: str) -> None:
    module = component.CodeModule
    if module.CountOfLines:
        module.DeleteLines(1, module.CountOfLines)
    module.AddFromString(code)


def format_start_sheet(workbook) -> None:
    sheet = workbook.Worksheets(1)
    sheet.Name = "START"
    sheet.Cells.Clear()

    sheet.Range("A1").Value = "OHADA COMPTA"
    sheet.Range("A2").Value = "Version Excel partageable — aucune DLL requise"
    sheet.Range("A4").Value = "Ce fichier .xlsm peut etre partage seul (aucun fichier supplementaire necessaire)."
    sheet.Range("A5").Value = "Si Excel affiche un bandeau de securite, faites clic droit > Proprietes > Debloquer, puis rouvrez le fichier."
    sheet.Range("A6").Value = "Activez les macros quand Excel le demande. Chrome ou Edge doit etre installe."
    sheet.Range("A8").Value = "Application chargee :"
    sheet.Range("A9").Value = LIVE_SITE_URL
    sheet.Range("A11").Value = "Aucune dependance externe : WebView2Loader.dll et les TLB ne sont pas necessaires."

    sheet.Range("A1").Font.Size = 22
    sheet.Range("A1").Font.Bold = True
    sheet.Range("A2").Font.Size = 12
    sheet.Range("A4:A11").Font.Size = 11
    sheet.Range("A9").Font.Color = 0xAA6C00
    sheet.Range("A9").Font.Underline = True
    sheet.Hyperlinks.Add(
        Anchor=sheet.Range("A9"),
        Address=LIVE_SITE_URL,
        TextToDisplay=LIVE_SITE_URL,
    )

    for column in ["A", "B", "C", "D", "E", "F"]:
        sheet.Columns(column).ColumnWidth = 26

    sheet.Range("A1:F20").Interior.Color = 0x0B110A
    sheet.Range("A1:F20").Font.Color = 0xE6E6E6
    sheet.Range("A1:F20").VerticalAlignment = -4160  # xlVAlignTop
    sheet.Range("A1:F20").HorizontalAlignment = -4131  # xlHAlignLeft
    sheet.Rows("1:20").RowHeight = 24
    sheet.Range("A1:F20").WrapText = True


def build_workbook(target_workbook: Path) -> None:
    target_workbook = target_workbook.resolve()
    target_workbook.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(TEMPLATE_WORKBOOK, target_workbook)

    pythoncom.CoInitialize()
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.AutomationSecurity = 3  # msoAutomationSecurityLow
    excel.EnableEvents = False

    try:
        workbook = excel.Workbooks.Open(str(target_workbook))
        try:
            excel.VBE.MainWindow.Visible = False
        except Exception:
            pass

        vb_project = workbook.VBProject

        # Drop the WebView2 COM reference — not needed in this build
        remove_reference_by_name(vb_project, "WebView2_edit")

        # Remove template components that are not used
        for component_name in UNUSED_COMPONENTS:
            remove_component_if_exists(vb_project, component_name)

        # Inject our hosted-browser VBA
        replace_component_code(vb_project.VBComponents("ThisWorkbook"), THISWORKBOOK_CODE)
        replace_component_code(vb_project.VBComponents("UserForm1"), USERFORM1_CODE)
        replace_component_code(ensure_standard_module(vb_project, "pluginExample"), HOST_MODULE_CODE)

        # Format the landing sheet
        format_start_sheet(workbook)

        workbook.Save()
        workbook.Close(SaveChanges=True)
    finally:
        excel.EnableEvents = True
        excel.Quit()
        pythoncom.CoUninitialize()


def build_fallback_workbook(target_path: Path) -> None:
    target_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "START"

    navy = "0B110A"
    gold = "AA6C00"
    light = "E6E6E6"

    for column in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[column].width = 26

    for row in range(1, 22):
        ws.row_dimensions[row].height = 24

    for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=6):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.font = Font(color=light, size=11)
            cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)

    ws["A1"] = "OHADA COMPTA EXACT"
    ws["A1"].font = Font(color=light, size=22, bold=True)
    ws["A2"] = "Version de secours sans macros"
    ws["A4"] = "Si le fichier .xlsm est bloque par Excel, utilisez ce classeur."
    ws["A5"] = "Cliquez sur le lien ci-dessous pour ouvrir OHADA Compta dans votre navigateur."
    ws["A7"] = LIVE_SITE_URL
    ws["A7"].hyperlink = LIVE_SITE_URL
    ws["A7"].font = Font(color=gold, size=12, bold=True, underline="single")
    ws["A9"] = "Cette edition ne depend d'aucune macro ni d'aucun composant externe."

    wb.save(target_path)


def build_zip_from_folder(folder: Path, zip_path: Path) -> None:
    if zip_path.exists():
        zip_path.unlink()

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for path in sorted(folder.rglob("*")):
            if path.is_file():
                zf.write(path, arcname=path.relative_to(folder.parent))


def build_install_pack(source_workbook: Path) -> None:
    if INSTALL_PACK_DIR.exists():
        shutil.rmtree(INSTALL_PACK_DIR)
    INSTALL_PACK_DIR.mkdir(parents=True, exist_ok=True)

    shutil.copy2(source_workbook, INSTALL_PACK_DIR / "OHADA-COMPTA-EXACT.xlsm")
    build_fallback_workbook(FALLBACK_XLSX)
    INSTALL_README.write_text(INSTALL_README_TEXT, encoding="utf-8")
    INSTALL_CMD.write_text(INSTALL_CMD_TEXT, encoding="utf-8")
    INSTALL_PS1.write_text(INSTALL_PS1_TEXT, encoding="utf-8")
    UNINSTALL_CMD.write_text(UNINSTALL_CMD_TEXT, encoding="utf-8")
    UNINSTALL_PS1.write_text(UNINSTALL_PS1_TEXT, encoding="utf-8")
    build_zip_from_folder(INSTALL_PACK_DIR, INSTALL_PACK_ZIP)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Build a single-file shareable OHADA-COMPTA-EXACT.xlsm that opens "
            "the live OHADA-Compta site inside Excel using a hosted Chrome/Edge "
            "window. No WebView2 DLL required."
        )
    )
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--desktop-output", type=Path, default=DEFAULT_DESKTOP_OUTPUT)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    print(f"Building {args.output} ...")
    build_workbook(args.output)
    print(f"BUILT: {args.output}")

    if args.desktop_output.resolve() != args.output.resolve():
        print(f"Building {args.desktop_output} ...")
        build_workbook(args.desktop_output)
        print(f"BUILT: {args.desktop_output}")

    build_install_pack(args.desktop_output)
    print(f"INSTALL_PACK: {INSTALL_PACK_DIR}")
    print(f"INSTALL_ZIP: {INSTALL_PACK_ZIP}")


if __name__ == "__main__":
    main()
