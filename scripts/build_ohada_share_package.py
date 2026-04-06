from __future__ import annotations

import shutil
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
DESKTOP = Path.home() / "OneDrive" / "Desktop"

SOURCE_WORKBOOK = ROOT / "OHADA-COMPTA-EXACT-SHAREABLE.xlsm"
PACKAGE_DIR = DESKTOP / "OHADA-COMPTA-SHARE-PACK"
PACKAGE_ZIP = DESKTOP / "OHADA-COMPTA-SHARE-PACK.zip"
FALLBACK_WORKBOOK = PACKAGE_DIR / "OHADA-COMPTA-WEB-LINK.xlsx"
README_FILE = PACKAGE_DIR / "README-IMPORTANT.txt"
LIVE_SITE_URL = "https://atarawendesidkabore-hash.github.io/ohada-compta/?v=20260406a"


README_TEXT = f"""OHADA COMPTA - PACK DE PARTAGE

Contenu:
- OHADA-COMPTA-EXACT-SHAREABLE.xlsm : version Excel qui embarque le vrai site dans Excel
- OHADA-COMPTA-WEB-LINK.xlsx : version de secours sans macros

Important:
Microsoft Excel bloque souvent les macros des fichiers recus par email, WhatsApp, Telegram ou telecharges depuis Internet.
Ce n'est pas un bug du fichier: c'est une protection de securite Windows/Excel.

Procedure recommandee pour le destinataire:
1. Si vous avez recu un fichier .zip, faites clic droit sur le .zip > Proprietes > Debloquer > Appliquer.
2. Extrayez ensuite le contenu du .zip.
3. Faites clic droit sur OHADA-COMPTA-EXACT-SHAREABLE.xlsm > Proprietes > Debloquer > Appliquer.
4. Ouvrez ensuite le fichier dans Excel Bureau.
5. Activez les macros si Excel le demande.
6. Verifiez que Google Chrome ou Microsoft Edge est installe sur le poste.

Si les macros restent bloquees:
- utilisez OHADA-COMPTA-WEB-LINK.xlsx
- ou ouvrez directement le site: {LIVE_SITE_URL}
"""


def build_fallback_workbook(target: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "START"

    navy = "061426"
    gold = "C8922A"
    light = "F2F2F2"

    for column in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[column].width = 26

    for row in range(1, 22):
        ws.row_dimensions[row].height = 24

    for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=6):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.font = Font(color=light, size=11)
            cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)

    ws["A1"] = "OHADA COMPTA"
    ws["A1"].font = Font(color=light, size=22, bold=True)
    ws["A2"] = "Version de secours sans macros"
    ws["A4"] = "Si le fichier .xlsm est bloque par la securite Excel, utilisez ce classeur."
    ws["A5"] = "Cliquez sur le lien ci-dessous pour ouvrir OHADA Compta dans votre navigateur."
    ws["A7"] = "Ouvrir OHADA Compta"
    ws["A7"].hyperlink = LIVE_SITE_URL
    ws["A7"].font = Font(color=gold, size=12, bold=True, underline="single")
    ws["A9"] = LIVE_SITE_URL
    ws["A9"].font = Font(color=gold, size=11, underline="single")
    ws["A11"] = "Conseil:"
    ws["A12"] = "Pour la version Excel complete, faites clic droit sur le fichier .xlsm > Proprietes > Debloquer, puis rouvrez-le."

    wb.save(target)


def build_zip_from_folder(folder: Path, zip_path: Path) -> None:
    if zip_path.exists():
      zip_path.unlink()

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for path in sorted(folder.rglob("*")):
            if path.is_file():
                zf.write(path, arcname=path.relative_to(folder.parent))


def main() -> None:
    if not SOURCE_WORKBOOK.exists():
        raise FileNotFoundError(f"Workbook introuvable: {SOURCE_WORKBOOK}")

    if PACKAGE_DIR.exists():
        shutil.rmtree(PACKAGE_DIR)
    PACKAGE_DIR.mkdir(parents=True, exist_ok=True)

    shutil.copy2(SOURCE_WORKBOOK, PACKAGE_DIR / SOURCE_WORKBOOK.name)
    README_FILE.write_text(README_TEXT, encoding="utf-8")
    build_fallback_workbook(FALLBACK_WORKBOOK)
    build_zip_from_folder(PACKAGE_DIR, PACKAGE_ZIP)

    print(f"PACKAGE_DIR:{PACKAGE_DIR}")
    print(f"PACKAGE_ZIP:{PACKAGE_ZIP}")


if __name__ == "__main__":
    main()
