from __future__ import annotations

import shutil
import zipfile
from pathlib import Path

import pythoncom
import win32com.client as win32
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
DESKTOP = Path.home() / "OneDrive" / "Desktop"

APP_LABEL = "OHADA COMPTA LITE"
LIVE_SITE_URL = "https://atarawendesidkabore-hash.github.io/ohada-compta/?v=20260405a"

LOCAL_WORKBOOK = ROOT / "OHADA-COMPTA-LITE-SHAREABLE.xlsm"
DESKTOP_WORKBOOK = DESKTOP / "OHADA-COMPTA-LITE-SHAREABLE.xlsm"
PACK_DIR = DESKTOP / "OHADA-COMPTA-LITE-SHARE-PACK"
PACK_ZIP = DESKTOP / "OHADA-COMPTA-LITE-SHARE-PACK.zip"
FALLBACK_XLSX = PACK_DIR / "OHADA-COMPTA-LITE-WEB-LINK.xlsx"
README_PATH = PACK_DIR / "README-IMPORTANT.txt"
PACK_WORKBOOK_NAME = "OHADA-COMPTA-LITE-SHAREABLE.xlsm"
INSTALL_PACK_DIR = DESKTOP / "OHADA-COMPTA-LITE-INSTALL-PACK"
INSTALL_PACK_ZIP = DESKTOP / "OHADA-COMPTA-LITE-INSTALL-PACK.zip"
INSTALL_README_PATH = INSTALL_PACK_DIR / "README-INSTALL.txt"
INSTALL_CMD_PATH = INSTALL_PACK_DIR / "INSTALL-OHADA-COMPTA-LITE.cmd"
INSTALL_PS1_PATH = INSTALL_PACK_DIR / "install-ohada-compta-lite.ps1"
UNINSTALL_CMD_PATH = INSTALL_PACK_DIR / "UNINSTALL-OHADA-COMPTA-LITE.cmd"
UNINSTALL_PS1_PATH = INSTALL_PACK_DIR / "uninstall-ohada-compta-lite.ps1"


LAUNCHER_MODULE_CODE = '''
Option Explicit

Public Const OHADA_COMPTA_URL As String = "{live_site_url}"

Public Sub LaunchOhadaCompta()
    If TryFollowHyperlink() Then Exit Sub
    If TryShellLaunch() Then Exit Sub
    ShowLaunchHelp
End Sub

Public Sub OpenInBrowser()
    LaunchOhadaCompta
End Sub

Private Function TryFollowHyperlink() As Boolean
    On Error GoTo Failed
    ThisWorkbook.FollowHyperlink Address:=OHADA_COMPTA_URL, NewWindow:=True
    TryFollowHyperlink = True
    Exit Function
Failed:
    TryFollowHyperlink = False
End Function

Private Function TryShellLaunch() As Boolean
    On Error GoTo Failed
    Dim shellApp As Object
    Set shellApp = CreateObject("WScript.Shell")
    shellApp.Run "cmd /c start """" """ & OHADA_COMPTA_URL & """", 0, False
    TryShellLaunch = True
    Exit Function
Failed:
    TryShellLaunch = False
End Function

Public Sub ShowLaunchHelp()
    MsgBox "Impossible d'ouvrir automatiquement OHADA Compta." & vbCrLf & vbCrLf & _
           "Ouvrez ce lien dans votre navigateur :" & vbCrLf & OHADA_COMPTA_URL, _
           vbInformation + vbOKOnly, "OHADA Compta Lite"
End Sub
'''.format(live_site_url=LIVE_SITE_URL).strip() + "\n"


THISWORKBOOK_CODE = """
Option Explicit

Private Sub Workbook_Open()
    On Error Resume Next
    Worksheets("START").Activate
    LiteShareLauncher.LaunchOhadaCompta
End Sub
""".strip() + "\n"


README_TEXT = f"""OHADA COMPTA LITE - PACK DE PARTAGE

Ce pack contient:
- OHADA-COMPTA-LITE-SHAREABLE.xlsm : version VBA ultra legere
- OHADA-COMPTA-LITE-WEB-LINK.xlsx : version sans macro

Pourquoi cette version est plus legere:
- elle n'embarque aucun navigateur dans Excel
- elle n'utilise aucune DLL ou controle WebView2
- elle ouvre simplement OHADA Compta dans le navigateur par defaut
- elle est adaptee aux ordinateurs plus anciens
- elle contient seulement un VBA minimal de lancement

Important:
Si le destinataire recoit un bandeau rouge de securite, ce n'est pas un crash du fichier.
Excel bloque souvent les macros des fichiers recus depuis Internet.

Procedure conseillee:
1. Envoyez de preference le fichier ZIP.
2. Le destinataire fait clic droit sur le ZIP > Proprietes > Debloquer > Appliquer.
3. Il extrait le ZIP.
4. Si besoin, il fait aussi clic droit sur OHADA-COMPTA-LITE-SHAREABLE.xlsm > Proprietes > Debloquer.
5. Il ouvre le fichier dans Excel.
6. Si rien ne s'ouvre, il clique sur le bouton "Ouvrir OHADA Compta".

Si les macros restent bloquees, utilisez directement:
{LIVE_SITE_URL}
ou le fichier OHADA-COMPTA-LITE-WEB-LINK.xlsx.
"""


INSTALL_README_TEXT = f"""OHADA COMPTA LITE - PACK D'INSTALLATION

Objectif:
Ce pack reduit le blocage "Potentially dangerous macro has been blocked" en installant le classeur
dans un emplacement de confiance Excel sur l'ordinateur du client.

Contenu:
- OHADA-COMPTA-LITE-SHAREABLE.xlsm
- OHADA-COMPTA-LITE-WEB-LINK.xlsx
- INSTALL-OHADA-COMPTA-LITE.cmd
- UNINSTALL-OHADA-COMPTA-LITE.cmd

Procedure recommandee:
1. Envoyez de preference ce pack au format ZIP.
2. Le client fait clic droit sur le ZIP > Proprietes > Debloquer > Appliquer.
3. Il extrait le ZIP.
4. Il lance INSTALL-OHADA-COMPTA-LITE.cmd.
5. Le script copie le classeur dans %LOCALAPPDATA%\\OHADA-Compta-Lite
   et cree un Trusted Location Excel pour les versions Office 2010/2013/2016/365.
6. Le client ouvre ensuite le raccourci Bureau "OHADA Compta Lite".

Ce que cela corrige:
- le classeur n'a plus besoin d'etre ouvert directement depuis Downloads, WhatsApp ou email
- le fichier est place dans un dossier local de confiance
- le script retire aussi le marquage Internet quand Windows l'autorise

Limites:
- ce n'est pas une signature numerique editeur
- certains postes tres verrouilles peuvent encore exiger une validation informatique
- si le client refuse tout script d'installation, utilisez la version sans macro:
  OHADA-COMPTA-LITE-WEB-LINK.xlsx

Site direct:
{LIVE_SITE_URL}
"""


INSTALL_CMD_TEXT = """@echo off
setlocal
cd /d "%~dp0"
powershell -NoProfile -ExecutionPolicy Bypass -File "%~dp0install-ohada-compta-lite.ps1"
if errorlevel 1 (
  echo.
  echo Installation terminee avec erreur.
  pause
  exit /b 1
)
echo.
echo Installation terminee.
pause
"""


UNINSTALL_CMD_TEXT = """@echo off
setlocal
cd /d "%~dp0"
powershell -NoProfile -ExecutionPolicy Bypass -File "%~dp0uninstall-ohada-compta-lite.ps1"
if errorlevel 1 (
  echo.
  echo Desinstallation terminee avec erreur.
  pause
  exit /b 1
)
echo.
echo Desinstallation terminee.
pause
"""


INSTALL_PS1_TEXT = r"""
$ErrorActionPreference = 'Stop'

$sourceRoot = Split-Path -Parent $MyInvocation.MyCommand.Path
$installRoot = Join-Path $env:LOCALAPPDATA 'OHADA-Compta-Lite'
$macroWorkbook = 'OHADA-COMPTA-LITE-SHAREABLE.xlsm'
$fallbackWorkbook = 'OHADA-COMPTA-LITE-WEB-LINK.xlsx'
$macroSource = Join-Path $sourceRoot $macroWorkbook
$fallbackSource = Join-Path $sourceRoot $fallbackWorkbook
$macroTarget = Join-Path $installRoot $macroWorkbook
$fallbackTarget = Join-Path $installRoot $fallbackWorkbook

if (-not (Test-Path $macroSource)) {
    throw "Fichier introuvable: $macroSource"
}

New-Item -ItemType Directory -Path $installRoot -Force | Out-Null
Copy-Item -Path $macroSource -Destination $macroTarget -Force
if (Test-Path $fallbackSource) {
    Copy-Item -Path $fallbackSource -Destination $fallbackTarget -Force
}

Get-ChildItem -Path $installRoot -File -ErrorAction SilentlyContinue | Unblock-File -ErrorAction SilentlyContinue

$trustedPathValue = $installRoot
if (-not $trustedPathValue.EndsWith('\')) {
    $trustedPathValue += '\'
}

foreach ($version in @('16.0', '15.0', '14.0')) {
    $trustedKey = "HKCU:\Software\Microsoft\Office\$version\Excel\Security\Trusted Locations\OHADAComptaLite"
    New-Item -Path $trustedKey -Force | Out-Null
    New-ItemProperty -Path $trustedKey -Name 'Path' -PropertyType String -Value $trustedPathValue -Force | Out-Null
    New-ItemProperty -Path $trustedKey -Name 'AllowSubfolders' -PropertyType DWord -Value 1 -Force | Out-Null
    New-ItemProperty -Path $trustedKey -Name 'Description' -PropertyType String -Value 'OHADA Compta Lite' -Force | Out-Null
    New-ItemProperty -Path $trustedKey -Name 'Date' -PropertyType String -Value (Get-Date -Format 'yyyy-MM-dd HH:mm:ss') -Force | Out-Null
}

$desktop = [Environment]::GetFolderPath('Desktop')
$shortcutPath = Join-Path $desktop 'OHADA Compta Lite.lnk'
$shell = New-Object -ComObject WScript.Shell
$shortcut = $shell.CreateShortcut($shortcutPath)
$shortcut.TargetPath = $macroTarget
$shortcut.WorkingDirectory = $installRoot
$shortcut.Description = 'OHADA Compta Lite'
$shortcut.IconLocation = "$env:SystemRoot\System32\shell32.dll,1"
$shortcut.Save()

Write-Host "INSTALL_ROOT=$installRoot"
Write-Host "SHORTCUT=$shortcutPath"
Write-Host "STATUS=OK"
""".strip() + "\n"


UNINSTALL_PS1_TEXT = r"""
$ErrorActionPreference = 'Stop'

$installRoot = Join-Path $env:LOCALAPPDATA 'OHADA-Compta-Lite'
$desktop = [Environment]::GetFolderPath('Desktop')
$shortcutPath = Join-Path $desktop 'OHADA Compta Lite.lnk'

foreach ($version in @('16.0', '15.0', '14.0')) {
    $trustedKey = "HKCU:\Software\Microsoft\Office\$version\Excel\Security\Trusted Locations\OHADAComptaLite"
    if (Test-Path $trustedKey) {
        Remove-Item -Path $trustedKey -Recurse -Force
    }
}

if (Test-Path $shortcutPath) {
    Remove-Item -Path $shortcutPath -Force
}

if (Test-Path $installRoot) {
    Remove-Item -Path $installRoot -Recurse -Force
}

Write-Host "STATUS=REMOVED"
""".strip() + "\n"


def ensure_standard_module(vb_project, name: str):
    try:
        return vb_project.VBComponents(name)
    except Exception:
        component = vb_project.VBComponents.Add(1)
        component.Name = name
        return component


def replace_component_code(component, code: str) -> None:
    module = component.CodeModule
    count = module.CountOfLines
    if count:
        module.DeleteLines(1, count)
    module.AddFromString(code)


def format_start_sheet(sheet, label: str, is_macro_version: bool) -> None:
    navy = 0x0B110A
    gold = 0xAA6C00

    sheet.Name = "START"
    sheet.Cells.Clear()

    sheet.Range("A1").Value = label
    sheet.Range("A2").Value = "Version Excel ultra legere"
    sheet.Range("A4").Value = "Cette edition est faite pour le partage et les vieux ordinateurs."
    sheet.Range("A5").Value = "Elle n'embarque pas le site dans Excel: elle ouvre simplement OHADA Compta dans le navigateur par defaut."
    sheet.Range("A7").Value = "Le lanceur essaie deux methodes d'ouverture pour mieux fonctionner sur des postes anciens."
    sheet.Range("A9").Value = LIVE_SITE_URL
    sheet.Range("A11").Value = "Action :"
    sheet.Range("A12").Value = "Ouvrir OHADA Compta dans le navigateur"
    sheet.Range("A14").Value = "Conseil partage : envoyez de preference le fichier ZIP, pas le .xlsm seul."
    if is_macro_version:
        sheet.Range("A16").Value = "Si rien ne se lance automatiquement, cliquez sur le bouton ci-dessous ou ouvrez le lien manuellement."
        sheet.Range("A18").Value = "Configuration minimale : Excel Bureau + un navigateur Windows."
    else:
        sheet.Range("A16").Value = "Cette version ne depend d'aucune macro."
        sheet.Range("A18").Value = "Utilisez cette edition si le .xlsm est bloque."

    sheet.Range("A1").Font.Size = 24
    sheet.Range("A1").Font.Bold = True
    sheet.Range("A2").Font.Size = 12
    sheet.Range("A9").Font.Color = gold
    sheet.Range("A9").Font.Underline = True
    sheet.Hyperlinks.Add(Anchor=sheet.Range("A9"), Address=LIVE_SITE_URL, TextToDisplay=LIVE_SITE_URL)

    sheet.Range("A1:F22").Interior.Color = navy
    sheet.Range("A1:F22").Font.Color = 0xE6E6E6
    sheet.Range("A1:F22").WrapText = True
    sheet.Range("A1:F22").VerticalAlignment = -4160
    sheet.Range("A1:F22").HorizontalAlignment = -4131

    for column in ["A", "B", "C", "D", "E", "F"]:
        sheet.Columns(column).ColumnWidth = 26

    sheet.Rows("1:22").RowHeight = 24

    if is_macro_version:
        button = sheet.Shapes.AddShape(1, 32, 360, 320, 44)
        button.TextFrame.Characters().Text = "Ouvrir OHADA Compta"
        button.Fill.ForeColor.RGB = 0x2AB8E5
        button.Line.ForeColor.RGB = 0x2AB8E5
        button.TextFrame.Characters().Font.Color = navy
        button.TextFrame.Characters().Font.Bold = True
        button.OnAction = "LiteShareLauncher.OpenInBrowser"


def build_macro_workbook(output_path: Path) -> None:
    output_path = output_path.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    pythoncom.CoInitialize()
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.AutomationSecurity = 3
    excel.EnableEvents = False

    try:
        workbook = excel.Workbooks.Add()
        workbook.SaveAs(str(output_path), FileFormat=52)
        vb_project = workbook.VBProject

        launcher_module = ensure_standard_module(vb_project, "LiteShareLauncher")
        replace_component_code(launcher_module, LAUNCHER_MODULE_CODE)
        replace_component_code(vb_project.VBComponents("ThisWorkbook"), THISWORKBOOK_CODE)
        format_start_sheet(workbook.Worksheets(1), APP_LABEL, True)

        workbook.Save()
        workbook.Close(SaveChanges=True)
    finally:
        excel.EnableEvents = True
        excel.Quit()
        pythoncom.CoUninitialize()


def build_fallback_workbook(output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    format_fallback_sheet(sheet)
    workbook.save(output_path)


def format_fallback_sheet(sheet) -> None:
    navy = "0B110A"
    gold = "AA6C00"
    light = "E6E6E6"

    sheet.title = "START"
    for column in ["A", "B", "C", "D", "E", "F"]:
        sheet.column_dimensions[column].width = 26

    for row in range(1, 23):
        sheet.row_dimensions[row].height = 24

    for row in sheet.iter_rows(min_row=1, max_row=22, min_col=1, max_col=6):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.font = Font(color=light, size=11)
            cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)

    sheet["A1"] = f"{APP_LABEL} WEB LINK"
    sheet["A1"].font = Font(color=light, size=24, bold=True)
    sheet["A2"] = "Version sans macro"
    sheet["A4"] = "Cette edition est la plus sure a partager."
    sheet["A5"] = "Elle ne lance rien automatiquement: cliquez simplement sur le lien ci-dessous."
    sheet["A7"] = LIVE_SITE_URL
    sheet["A7"].hyperlink = LIVE_SITE_URL
    sheet["A7"].font = Font(color=gold, size=12, bold=True, underline="single")
    sheet["A9"] = "A utiliser si Excel bloque les macros du fichier .xlsm."
    sheet["A11"] = "Cette version est recommandee pour les environnements les plus limites."


def build_zip_from_folder(folder: Path, zip_path: Path) -> None:
    if zip_path.exists():
        zip_path.unlink()

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for path in sorted(folder.rglob("*")):
            if path.is_file():
                zf.write(path, arcname=path.relative_to(folder.parent))


def build_share_pack(source_macro_workbook: Path) -> None:
    if PACK_DIR.exists():
        shutil.rmtree(PACK_DIR)
    PACK_DIR.mkdir(parents=True, exist_ok=True)

    shutil.copy2(source_macro_workbook, PACK_DIR / PACK_WORKBOOK_NAME)
    README_PATH.write_text(README_TEXT, encoding="utf-8")
    build_fallback_workbook(FALLBACK_XLSX)
    build_zip_from_folder(PACK_DIR, PACK_ZIP)


def build_install_pack(source_macro_workbook: Path) -> None:
    if INSTALL_PACK_DIR.exists():
        shutil.rmtree(INSTALL_PACK_DIR)
    INSTALL_PACK_DIR.mkdir(parents=True, exist_ok=True)

    shutil.copy2(source_macro_workbook, INSTALL_PACK_DIR / PACK_WORKBOOK_NAME)
    build_fallback_workbook(INSTALL_PACK_DIR / "OHADA-COMPTA-LITE-WEB-LINK.xlsx")
    INSTALL_README_PATH.write_text(INSTALL_README_TEXT, encoding="utf-8")
    INSTALL_CMD_PATH.write_text(INSTALL_CMD_TEXT, encoding="utf-8")
    INSTALL_PS1_PATH.write_text(INSTALL_PS1_TEXT, encoding="utf-8")
    UNINSTALL_CMD_PATH.write_text(UNINSTALL_CMD_TEXT, encoding="utf-8")
    UNINSTALL_PS1_PATH.write_text(UNINSTALL_PS1_TEXT, encoding="utf-8")
    build_zip_from_folder(INSTALL_PACK_DIR, INSTALL_PACK_ZIP)


def main() -> None:
    build_macro_workbook(LOCAL_WORKBOOK)

    desktop_target = DESKTOP_WORKBOOK
    if desktop_target.exists():
        desktop_target = DESKTOP / "OHADA-COMPTA-LITE-SHAREABLE-v2.xlsm"
    build_macro_workbook(desktop_target)

    build_share_pack(desktop_target)
    build_install_pack(desktop_target)

    print(f"BUILT:{LOCAL_WORKBOOK}")
    print(f"BUILT:{desktop_target}")
    print(f"PACK:{PACK_DIR}")
    print(f"ZIP:{PACK_ZIP}")
    print(f"INSTALL_PACK:{INSTALL_PACK_DIR}")
    print(f"INSTALL_ZIP:{INSTALL_PACK_ZIP}")


if __name__ == "__main__":
    main()
